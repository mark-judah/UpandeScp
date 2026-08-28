"""The weekly block report for block-grown crops — avocado and coffee.

Roses are scouted bed by bed inside greenhouses, and their weekly submission is the
six-sheet KEPHIS FCM workbook: a regulated template, specific to false codling moth.
Avocado and coffee are grown on **blocks**, are not FCM-reportable, and need something
much plainer — one sheet, blocks down the rows, pests across the columns, and the
week's counts in the cells.

## What a cell holds

`SUM(count)` from `Pests Scouting Entry` for that pest, in that block, over the ISO
week. The child row records `plant_section`, `pest`, `stage` and `count`; summing count
is the literal "cumulative over that week" and stays comparable between blocks of
similar size. Stage and plant section are summed over rather than broken out — a
sheet with a column per (pest, stage, section) is unreadable at avocado's pest count.

Pests only. Diseases, weeds, predators and disorders are recorded by the same scouts on
the same entries, but this sheet answers one question.

## Why a block with no data still gets a row

A block that was walked and found clean and a block nobody visited are different facts,
and a sheet that omits both makes them identical. Every block on the farm gets a row;
an unvisited one is visibly all-zero rather than absent.

## Why a farm with no blocks is refused rather than emptied

Coffee is tagged to Endebess and Saboti, and neither has a single warehouse typed as a
`Block` — 96 and 23 warehouses between them, all untyped. A report for those farms is
not empty because nothing was found; it is empty because there is nothing to look at.
`availability()` says so, and the page refuses the download rather than handing over a
file whose blankness means something different from what it looks like.
"""

from __future__ import annotations

import io
from collections import defaultdict

import frappe

from upande_scp.serverscripts.common import crop_scope
from upande_scp.serverscripts.scouting.get_complete_scouting_entries import _week_bounds

#: Warehouses of this type are the reporting unit for block-grown crops.
BLOCK_TYPE = "Block"


def _blocks_for_farm(farm: str) -> list[str]:
	rows = frappe.get_all(
		"Warehouse",
		filters={
			"custom_farm": farm,
			"warehouse_type": BLOCK_TYPE,
			"disabled": 0,
			"is_group": 0,
		},
		fields=["name"],
		order_by="name",
	)
	return [r["name"] for r in rows]


@frappe.whitelist()
def availability(crop: str) -> dict:
	"""Which of this crop's farms can produce a report, and why the others cannot.

	Returned before any download so the page can explain a farm with no blocks rather
	than offer a file that looks empty for the wrong reason.
	"""
	if not crop:
		frappe.throw("A crop is required.")

	farms = crop_scope.scoped_farms(crop, frappe.session.user)
	if farms is None:
		farms = crop_scope.farms_for_crop(crop) or set()

	ready, blocked = [], []
	for farm in sorted(farms):
		blocks = _blocks_for_farm(farm)
		if blocks:
			ready.append({"farm": farm, "blocks": len(blocks)})
		else:
			blocked.append({
				"farm": farm,
				"reason": (
					f"{farm} has no blocks set up yet — its warehouses are not typed as "
					f"'{BLOCK_TYPE}', so there is nothing to report on."
				),
			})

	return {"crop": crop, "ready": ready, "blocked": blocked}


@frappe.whitelist()
def report_weeks(crop: str, farm: str, limit: int = 26) -> list[dict]:
	"""ISO weeks that actually have scouting on this farm, newest first.

	Offered instead of a free week picker so nobody downloads a blank sheet for a week
	nobody walked — the same reason `availability` exists one level up.
	"""
	blocks = _blocks_for_farm(farm)
	if not blocks:
		return []
	rows = frappe.db.sql(
		"""
		SELECT YEAR(se.date_of_capture) AS iso_year,
		       WEEK(se.date_of_capture, 3) AS iso_week,
		       COUNT(*) AS entries
		FROM `tabScouting Entry` se
		WHERE se.block IN %(blocks)s AND se.crop_scouted = %(crop)s
		  AND se.date_of_capture IS NOT NULL
		GROUP BY iso_year, iso_week
		ORDER BY iso_year DESC, iso_week DESC
		LIMIT %(limit)s
		""",
		{"blocks": tuple(blocks), "crop": crop, "limit": int(limit)},
		as_dict=True,
	)
	out = []
	for r in rows:
		monday, sunday = _week_bounds(int(r["iso_year"]), int(r["iso_week"]))
		out.append({
			"year": int(r["iso_year"]),
			"week": int(r["iso_week"]),
			"entries": int(r["entries"]),
			"label": f"W{int(r['iso_week']):02d} · {monday} to {sunday}",
		})
	return out


def _pest_counts(blocks: list[str], monday, sunday, crop: str) -> dict:
	"""``{block: {pest: total}}`` for one ISO week.

	One query rather than one per block: a farm can carry a hundred blocks, and this
	runs behind a download the user is waiting on.
	"""
	if not blocks:
		return {}
	# `block`, not `greenhouse`. A block-grown crop records its unit in `Scouting
	# Entry.block`; `greenhouse` is NULL on every one of the 3,362 avocado entries.
	# Joining on the wrong one returns an empty sheet that looks like a clean week.
	rows = frappe.db.sql(
		"""
		SELECT se.block AS block, p.pest AS pest, SUM(p.count) AS total
		FROM `tabPests Scouting Entry` p
		JOIN `tabScouting Entry` se ON se.name = p.parent
		WHERE se.block IN %(blocks)s
		  AND se.crop_scouted = %(crop)s
		  AND se.date_of_capture BETWEEN %(monday)s AND %(sunday)s
		  AND p.pest IS NOT NULL AND p.pest != ''
		GROUP BY se.block, p.pest
		""",
		{
			"blocks": tuple(blocks),
			"crop": crop,
			"monday": monday,
			"sunday": sunday,
		},
		as_dict=True,
	)
	out: dict = defaultdict(dict)
	for r in rows:
		out[r["block"]][r["pest"]] = float(r["total"] or 0)
	return out


def _pests_for_crop(crop: str) -> list[str]:
	"""Every pest this crop's filters name, so the columns are stable week to week.

	Taken from the crop's `Pest Filter` records rather than from the week's data: a
	pest that was not seen this week still deserves a column, or the sheet's shape
	changes every week and cannot be compared or pivoted.
	"""
	rows = frappe.get_all(
		"Pest Filter", filters={"crop_scouted": crop}, fields=["pest"], order_by="pest"
	)
	pests = []
	for r in rows:
		if r.get("pest") and r["pest"] not in pests:
			pests.append(r["pest"])
	return pests


def build_workbook_bytes(crop: str, farm: str, iso_year: int, iso_week: int) -> bytes:
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill

	blocks = _blocks_for_farm(farm)
	if not blocks:
		frappe.throw(
			f"{farm} has no blocks set up, so there is nothing to report on.",
			frappe.ValidationError,
		)

	monday, sunday = _week_bounds(iso_year, iso_week)
	counts = _pest_counts(blocks, monday, sunday, crop)

	pests = _pests_for_crop(crop)
	# A pest seen this week but absent from the crop's filters still gets a column —
	# the data is the authority on what was found, the filters only on what to expect.
	for block_counts in counts.values():
		for pest in block_counts:
			if pest not in pests:
				pests.append(pest)

	wb = Workbook()
	ws = wb.active
	ws.title = f"{crop} W{iso_week:02d}"

	ws["A1"] = f"{crop} — weekly pest counts"
	ws["A1"].font = Font(bold=True, size=14)
	ws["A2"] = f"{farm}   ·   {iso_year}-W{iso_week:02d}   ·   {monday} to {sunday}"
	ws["A3"] = "Each cell is the total count recorded for that pest on that block over the week."

	header_row = 5
	ws.cell(row=header_row, column=1, value="Block")
	for i, pest in enumerate(pests):
		ws.cell(row=header_row, column=2 + i, value=pest)
	ws.cell(row=header_row, column=2 + len(pests), value="Total")

	head_fill = PatternFill("solid", fgColor="DDDDDD")
	for c in range(1, 3 + len(pests)):
		cell = ws.cell(row=header_row, column=c)
		cell.font = Font(bold=True)
		cell.fill = head_fill
		cell.alignment = Alignment(horizontal="center", wrap_text=True)

	column_totals = [0.0] * len(pests)
	for r, block in enumerate(blocks, start=header_row + 1):
		ws.cell(row=r, column=1, value=block)
		row_total = 0.0
		for i, pest in enumerate(pests):
			value = counts.get(block, {}).get(pest, 0)
			ws.cell(row=r, column=2 + i, value=value)
			row_total += value
			column_totals[i] += value
		ws.cell(row=r, column=2 + len(pests), value=row_total)

	total_row = header_row + 1 + len(blocks)
	ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
	for i, total in enumerate(column_totals):
		cell = ws.cell(row=total_row, column=2 + i, value=total)
		cell.font = Font(bold=True)
	grand = ws.cell(row=total_row, column=2 + len(pests), value=sum(column_totals))
	grand.font = Font(bold=True)

	ws.column_dimensions["A"].width = 32
	for i in range(len(pests) + 1):
		ws.column_dimensions[ws.cell(row=header_row, column=2 + i).column_letter].width = 14
	ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


@frappe.whitelist()
def download_block_weekly_xlsx(crop: str, farm: str, week=None, year=None):
	from datetime import date

	crop_scope.assert_crop(crop)
	allowed = crop_scope.scoped_farms(crop, frappe.session.user)
	if allowed is not None and farm not in allowed:
		frappe.throw(
			f"{farm} is not a farm you have access to for {crop}.",
			frappe.PermissionError,
		)

	today = date.today()
	iso_year = int(year) if year else today.isocalendar()[0]
	iso_week = int(week) if week else today.isocalendar()[1]

	data = build_workbook_bytes(crop, farm, iso_year, iso_week)
	safe_farm = str(farm).replace(" ", "_").replace("/", "-")
	frappe.local.response.filename = (
		f"{crop}_{safe_farm}_{iso_year}-W{iso_week:02d}.xlsx"
	)
	frappe.local.response.filecontent = data
	frappe.local.response.type = "binary"
