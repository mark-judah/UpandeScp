"""
send_fcm_weekly_excel_report.py
================================
Generate and email the KEPHIS FCM Weekly Monitoring Excel report.
Mirrors the format of the official KEPHIS template.

One workbook is produced **per farm**. Farm identity (display name,
KEPHIS site id, abbreviation) is sourced from the Farm doctype — nothing
is hard-coded here. The scheduler loops over every farm that has
scouting data for the year and is enabled in Scouting and Crop Protection Settings.

Sheets generated
----------------
1. FCM Daily monitoring   – per-trap FCM counts (indoor / outdoor) by week
2. Weekly summary         – all moth species totals by week
3. Scouting Summary       – plant-level pests per GH per week
4. Intake QC Report       – (empty — pending AI integration)
5. Variety List           – (empty — pending integration with live variety data)
6. FCM Risk profiling     – (empty — pending AI integration)

The per-observation audit trail is NOT in this workbook. It is a separate,
week-scoped download — see ``download_scouting_entries_xlsx``. It used to be a
7th sheet here carrying every observation from week 1 to the selected week,
which on a large farm late in the year meant ~350,000 individually-styled rows
and about two minutes of openpyxl, overrunning the request timeout and taking
the KEPHIS submission down with it.

Hook in hooks.py:
    scheduler_events = {
        "weekly": [
            "upande_scp.serverscripts.reports.send_fcm_weekly_excel_report.send_fcm_weekly_excel_report"
        ]
    }
"""

import io
import re
from datetime import date, timedelta

import frappe

from upande_scp.serverscripts.scouting.scouting_metrics import (
    MOTH_OTHERS,
    get_fcm_larvae_weekly,
    get_fcm_trap_counts_weekly,
    get_fcm_traps_ordered,
    get_plant_pests_weekly,
    get_scouting_records_weekly,
    get_weekly_trap_pest_totals_indoor,
)
from upande_scp.upande_scp.doctype.scouting_and_crop_protection_settings.scouting_and_crop_protection_settings import (
    get_allowed_farms,
)


# ---------------------------------------------------------------------------
# Farm resolution
# ---------------------------------------------------------------------------

def _resolve_farm(farm_name):
    """Return (display_name, kephis_id, abbreviation) for a farm.

    Falls back gracefully when the Farm doctype is absent or the row is
    missing — the report should never crash just because metadata is
    incomplete.
    """
    display = farm_name or ""
    kephis  = ""
    abbrev  = ""
    try:
        if farm_name and frappe.db.exists("Farm", farm_name):
            row = frappe.db.get_value(
                "Farm", farm_name,
                ["farm_name", "farm_code", "abbreviation"],
                as_dict=True,
            ) or {}
            display = row.get("farm_name") or farm_name
            kephis  = row.get("farm_code") or ""
            abbrev  = row.get("abbreviation") or ""
    except Exception:
        # Farm doctype not installed or query failed — keep defaults.
        pass
    return display, kephis, abbrev


def _farms_with_data(year):
    """Farms that have at least one scouting entry in ``year`` AND are
    enabled in Scouting and Crop Protection Settings.

    Uses ``Warehouse.custom_farm`` as the authoritative link from a
    scouting entry's greenhouse to a farm. When Scouting and Crop Protection Settings has
    no allowed farms configured, every farm with scouting data is
    returned so existing setups keep working.
    """
    rows = frappe.db.sql(
        """
        SELECT DISTINCT gh.custom_farm AS farm
        FROM  `tabScouting Entry` se
        JOIN  `tabWarehouse` gh ON gh.name = se.greenhouse
        WHERE YEAR(se.date_of_capture) = %s
          AND gh.custom_farm IS NOT NULL
          AND gh.custom_farm != ''
        ORDER BY gh.custom_farm
        """,
        (year,),
        as_dict=True,
    )
    farms = [r.farm for r in rows]
    allowed = set(get_allowed_farms())
    if allowed:
        farms = [f for f in farms if f in allowed]
    return farms


# ---------------------------------------------------------------------------
# Internal: workbook builder (shared by email and download)
# ---------------------------------------------------------------------------

def _build_workbook_bytes(farm, week=None, year=None):
    """Build the KEPHIS FCM Excel workbook for a single farm.

    ``week`` selects the ISO week (1-53) the report targets — the week
    highlighted green and named in the filename, with data shown cumulatively
    from the start of the year through it. When omitted, the last completed
    week is used (original scheduler behaviour).

    Returns (excel_bytes, fname, current_year, week_range_str, last_week_num,
             farm_display).
    When there's no data for the year/farm, excel_bytes is None.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.log_error("openpyxl not installed – run: pip install openpyxl", "FCM Weekly Report")
        return None, None, None, None, None, None

    farm_display, kephis_id, abbrev = _resolve_farm(farm)

    today = date.today()
    current_year = int(year) if year else today.year
    weekday = today.weekday()          # 0 = Monday
    current_week_num = int(today.isocalendar()[1])

    # The report targets one ISO week (Mon → Sun). Honour the caller's choice
    # when given, otherwise default to the last completed week.
    default_week_mon = today - timedelta(days=weekday + 7)
    default_week_num = int(default_week_mon.isocalendar()[1])
    report_week_num = int(week) if week else default_week_num

    # ``last_week_num`` keeps its original name (used for the green highlight
    # and the filename) but now points at the selected report week.
    last_week_num = report_week_num

    try:
        rw_mon = date.fromisocalendar(current_year, report_week_num, 1)
        rw_sun = date.fromisocalendar(current_year, report_week_num, 7)
    except ValueError:
        rw_mon = default_week_mon
        rw_sun = default_week_mon + timedelta(days=6)

    week_range_str = (
        rw_mon.strftime("%d %b") + " - " + rw_sun.strftime("%d %b %Y")
    )

    # -----------------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------------

    def parse_trap_num(trap_number):
        try:
            return int(str(trap_number).strip())
        except Exception:
            return 0

    def gh_from_trap(trap_number):
        """Derive greenhouse number from trap number (e.g. 1307 → 13)."""
        n = parse_trap_num(trap_number)
        return n // 100 if n >= 100 else 0

    def gh_from_warehouse(gh_name):
        """Extract GH number from a warehouse name like 'X GH 12 - KR'."""
        if not gh_name:
            return 0
        m = re.search(r"GH\s+(\d+)", str(gh_name))
        return int(m.group(1)) if m else 0

    def safe_int(v):
        try:
            return int(float(v or 0))
        except Exception:
            return 0

    def iso_week_range(year, week_num):
        mon = date.fromisocalendar(year, week_num, 1)
        sun = date.fromisocalendar(year, week_num, 7)
        return mon, sun

    def human_date(d):
        """Format date as '5th Jan 2026'."""
        day = d.day
        sfx = "th" if 11 <= (day % 100) <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        return f"{day}{sfx} {d.strftime('%b %Y')}"

    def map_stage(stage):
        s = (stage or "").lower()
        if "egg" in s:
            return "eggs"
        if "larva" in s or "nymph" in s or "instar" in s:
            return "larvae"
        if "damage" in s:
            return "damages"
        return "other"

    # -----------------------------------------------------------------------
    # Data (sourced from scouting_metrics, scoped to this farm)
    # -----------------------------------------------------------------------

    fcm_traps      = get_fcm_traps_ordered(farm=farm)
    fcm_raw        = get_fcm_trap_counts_weekly(current_year, farm=farm)
    trap_pest_wk   = get_weekly_trap_pest_totals_indoor(current_year, farm=farm)
    plant_pests    = get_plant_pests_weekly(current_year, farm=farm)
    fcm_larvae_raw = get_fcm_larvae_weekly(current_year, farm=farm)

    # -----------------------------------------------------------------------
    # Build lookup maps
    # -----------------------------------------------------------------------

    # {(wk, trap_name, 'Indoor'|'Outdoor'): count}
    fcm_trap_map = {}
    for r in fcm_raw:
        loc = "Outdoor" if "out" in (r.location or "").lower() else "Indoor"
        key = (r.wk, r.trap, loc)
        fcm_trap_map[key] = fcm_trap_map.get(key, 0) + safe_int(r.cnt)

    # {(wk, pest): count}
    trap_pest_map = {}
    for r in trap_pest_wk:
        pest = (r.pest or "").strip()
        trap_pest_map[(r.wk, pest)] = trap_pest_map.get((r.wk, pest), 0) + safe_int(r.cnt)

    # {(wk, gh_num): larvae_count}
    fcm_larvae_map = {}
    for r in fcm_larvae_raw:
        gh = gh_from_warehouse(r.greenhouse)
        key = (r.wk, gh)
        fcm_larvae_map[key] = fcm_larvae_map.get(key, 0) + safe_int(r.cnt)

    # {(wk, gh_num, pest_cat, stage_cat): count}   pest_cat ∈ {FCM, Helicoverpa, Others}
    # "Others" is restricted to moth species only (MOTH_OTHERS) — non-moth pests
    # like aphids/thrips/mites are excluded from this sheet entirely.
    scouting_map = {}
    for r in plant_pests:
        pest = (r.pest or "").strip()
        if pest == "FCM":
            p_cat = "FCM"
        elif pest == "Helicoverpa":
            p_cat = "Helicoverpa"
        elif pest in MOTH_OTHERS:
            p_cat = "Others"
        else:
            continue
        gh = gh_from_warehouse(r.greenhouse)
        s_cat = map_stage(r.stage)
        key = (r.wk, gh, p_cat, s_cat)
        scouting_map[key] = scouting_map.get(key, 0) + safe_int(r.cnt)

    # Collect all weeks that have any data, capped at current week
    all_wk_nums = set()
    for r in fcm_raw:
        all_wk_nums.add(r.wk)
    for r in trap_pest_wk:
        all_wk_nums.add(r.wk)
    for r in plant_pests:
        all_wk_nums.add(r.wk)

    all_weeks = sorted(w for w in all_wk_nums if 1 <= w <= report_week_num)

    if not all_weeks:
        return None, None, current_year, week_range_str, last_week_num, farm_display

    # Sorted trap list
    sorted_traps = sorted(fcm_traps, key=lambda t: parse_trap_num(t.get("trap_number") or "0"))

    # Dynamic greenhouse range — derived from the traps actually configured for
    # this farm, not hard-coded. Fallback to union of GHs that appear in
    # scouting data for the Scouting Summary sheet.
    gh_nums = set()
    for t in sorted_traps:
        n = gh_from_trap(t.get("trap_number") or "0")
        if n:
            gh_nums.add(n)
    for r in plant_pests:
        n = gh_from_warehouse(r.greenhouse)
        if n:
            gh_nums.add(n)
    for r in fcm_larvae_raw:
        n = gh_from_warehouse(r.greenhouse)
        if n:
            gh_nums.add(n)
    GHS = sorted(gh_nums) if gh_nums else []

    # -----------------------------------------------------------------------
    # Style factories
    # -----------------------------------------------------------------------

    def mk_font(bold=False, color="000000", size=11):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def mk_fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def mk_border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    DARK_BLUE = "0D2B5E"
    LIGHT_BLUE = "DDE6F5"
    GREEN_HL = "C6EFCE"
    DARK_GREEN = "538135"
    MID_BLUE = "1F5C99"

    dark_fill  = mk_fill(DARK_BLUE)
    light_fill = mk_fill(LIGHT_BLUE)
    green_fill = mk_fill(GREEN_HL)

    hdr_fnt  = mk_font(bold=True, color="FFFFFF")
    bold_fnt = mk_font(bold=True)
    data_fnt = mk_font()
    italic_note_fnt = Font(name="Calibri", italic=True, color="666666", size=11)

    thin_brd   = mk_border()
    c_align    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align    = Alignment(horizontal="left",   vertical="center")
    r_align    = Alignment(horizontal="right",  vertical="center")

    farm_title = (farm_display or "").upper() or "-"

    # -----------------------------------------------------------------------
    # Create workbook
    # -----------------------------------------------------------------------
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===================================================================
    # SHEET 1 – FCM Daily monitoring
    # ===================================================================
    ws1 = wb.create_sheet("FCM Daily monitoring")
    for i, w in enumerate([20, 30, 12, 24, 24, 28, 20, 34, 20], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Title block
    ws1["A1"] = "Production Site Name"; ws1["A1"].font = bold_fnt
    ws1["E1"] = "Year";                 ws1["E1"].font = bold_fnt
    ws1["F1"] = current_year
    ws1["A2"] = farm_title;             ws1["A2"].font = mk_font(bold=True, size=14)
    if kephis_id:
        ws1["C2"] = f"KEPHIS ID: {kephis_id}"
        ws1["C2"].font = bold_fnt
    ws1["A3"] = "Share the FCM data via email rosafcmdata@kephis.org"
    ws1["A4"] = "Instructions: "
    ws1["A5"] = "(a) The number of traps depends on the size of the farm at a density of 4 traps per Ha."
    ws1["A6"] = "(b) Traps to be placed strategically outside greenhouse at an interval of 50 m"
    ws1["A7"] = "(c) Data to be collected daily and cumulative counts reported to KEPHIS every Monday"
    ws1["A8"] = "(d) Ensure instructions are followed when placing/servicing FCM pheromone lures and delta traps"
    ws1["A9"] = "Fill all the other sheets accordingly"

    ws1.merge_cells("C11:I11")
    ws1["C11"] = "FCM COUNTS"
    ws1["C11"].font = hdr_fnt
    ws1["C11"].fill = dark_fill
    ws1["C11"].alignment = c_align

    _h1 = [
        "Week of the year",
        "Greenhouse No./Identity and size (Ha)",
        "Trap No.",
        "Source (supplier) of pheromone lure",
        "Pheromone placement date",
        "Date of count",
        "No. of FCM adults inside greenhouse",
        "Cummulative No. of eggs/larvae observed in the greenhouse",
        "No. of FCM adults outside greenhouse",
    ]
    ws1.row_dimensions[12].height = 44
    for col, hdr in enumerate(_h1, 1):
        c = ws1.cell(row=12, column=col, value=hdr)
        c.font = hdr_fnt; c.fill = dark_fill; c.border = thin_brd; c.alignment = c_align

    cur1 = 13
    for week_num in all_weeks:
        try:
            wk_mon, wk_sun = iso_week_range(current_year, week_num)
        except Exception:
            continue

        is_last = (week_num == last_week_num)
        w_fill  = green_fill if is_last else None

        wk_label = f"Week {week_num:02d}"
        from_str = f"FROM: {human_date(wk_mon)}"
        to_str   = f"TO: {human_date(wk_sun)}"

        row_in_wk = 0
        last_gh   = None
        wk_start  = cur1

        for trap in sorted_traps:
            trap_name  = trap.get("name") or ""
            trap_num_s = str(trap.get("trap_number") or "")
            trap_gh    = gh_from_trap(trap_num_s)
            trap_loc   = (trap.get("location") or "Indoor").strip()

            indoor_cnt  = fcm_trap_map.get((week_num, trap_name, "Indoor"),  0)
            outdoor_cnt = fcm_trap_map.get((week_num, trap_name, "Outdoor"), 0)

            # Assign counts to correct column based on trap's registered location
            if trap_loc == "Outdoor":
                col_g_val = None
                col_i_val = indoor_cnt + outdoor_cnt
            else:
                col_g_val = indoor_cnt + outdoor_cnt
                col_i_val = None

            # Cumulative larvae: only on first trap of each GH group
            is_first_gh = (trap_gh != last_gh)
            col_h_val   = fcm_larvae_map.get((week_num, trap_gh), 0) if is_first_gh else None

            # Column A labels (week / from / to on first 3 rows of week)
            col_a = wk_label if row_in_wk == 0 else (from_str if row_in_wk == 1 else (to_str if row_in_wk == 2 else None))
            # Column B: house label on first trap of each GH
            col_b = f"House {trap_gh:02d}" if is_first_gh else None
            # Column F: date range on first two rows
            col_f = from_str if row_in_wk == 0 else (to_str if row_in_wk == 1 else None)

            vals = [col_a, col_b, trap_num_s, "KOPPERT", wk_mon, col_f, col_g_val, col_h_val, col_i_val]
            for col, val in enumerate(vals, 1):
                c = ws1.cell(row=cur1, column=col, value=val)
                c.font  = data_fnt
                c.border = thin_brd
                if w_fill:
                    c.fill = w_fill
                c.alignment = r_align if col in (7, 8, 9) else l_align

            last_gh = trap_gh
            cur1 += 1
            row_in_wk += 1

        # Week total row
        wk_end = cur1 - 1
        for col in range(1, 10):
            c = ws1.cell(row=cur1, column=col)
            c.font  = mk_font(bold=True, color="FFFFFF")
            c.fill  = dark_fill
            c.border = thin_brd
            c.alignment = r_align
        ws1.cell(row=cur1, column=1).value = "WEEK TOTAL"
        ws1.cell(row=cur1, column=1).alignment = l_align
        if wk_end >= wk_start:
            ws1.cell(row=cur1, column=7).value = f"=SUM(G{wk_start}:G{wk_end})"
            ws1.cell(row=cur1, column=8).value = f"=SUM(H{wk_start}:H{wk_end})"
            ws1.cell(row=cur1, column=9).value = f"=SUM(I{wk_start}:I{wk_end})"
        cur1 += 2  # blank separator row between weeks

    # ===================================================================
    # SHEET 2 – Weekly summary
    # ===================================================================
    ws2 = wb.create_sheet("Weekly summary")
    for i, w in enumerate([14, 12, 18, 20, 20, 22, 22, 22, 22, 14, 10, 10, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2["A1"] = "Production Site Name"; ws2["A1"].font = bold_fnt
    ws2["E1"] = "Year";                 ws2["E1"].font = bold_fnt
    ws2["F1"] = current_year
    ws2["A2"] = farm_title;             ws2["A2"].font = mk_font(bold=True, size=14)

    ws2.merge_cells("A3:M3")
    ws2["A3"] = f"LIGHT AND PHEROMONE TRAPS INSECT/MOTH TOTAL COUNTS SUMMARY/ANALYSIS {current_year}"
    ws2["A3"].font = mk_font(bold=True, size=12)
    ws2["A3"].alignment = c_align

    _h2 = [
        "Month of count", "Week of the year",
        "FCM adults\n(light)", "FCM adults\n(pheromone)", "New Cases of\nFCM Adults",
        "Helicoverpa adults\n(light trap)", "Spodoptera adults\n(light trap)",
        "Helicoverpa adults\n(Lure Trap)", "Spodoptera adults\n(Lure Trap)",
        "Others (light)",
        "Temperature (°C)", None, "Rainfall (mm)",
    ]
    ws2.row_dimensions[4].height = 48
    ws2.row_dimensions[5].height = 16
    for col, hdr in enumerate(_h2, 1):
        if hdr is None:
            continue
        c = ws2.cell(row=4, column=col, value=hdr)
        c.font = hdr_fnt; c.fill = dark_fill; c.border = thin_brd; c.alignment = c_align

    ws2.merge_cells("K4:L4")
    for sub, col in [("Min.", 11), ("Max.", 12)]:
        c = ws2.cell(row=5, column=col, value=sub)
        c.font = hdr_fnt; c.fill = dark_fill; c.border = thin_brd; c.alignment = c_align

    cur2 = 6
    last_month = None
    for week_num in all_weeks:
        try:
            wk_mon, _ = iso_week_range(current_year, week_num)
        except Exception:
            continue

        is_last = (week_num == last_week_num)
        w_fill2  = green_fill if is_last else None

        month     = wk_mon.strftime("%B").upper()
        month_val = month if month != last_month else None
        last_month = month

        fcm_p   = trap_pest_map.get((week_num, "FCM"), 0)         or None
        hel_l   = trap_pest_map.get((week_num, "Helicoverpa"), 0) or None
        spo_l   = trap_pest_map.get((week_num, "Spodoptera"), 0)  or None
        dup     = trap_pest_map.get((week_num, "Duponchella"), 0)
        unid    = trap_pest_map.get((week_num, "Unidentified Moth"), 0)
        others  = (dup + unid) or None

        row_data2 = [
            month_val, f"WK {week_num:02d}",
            None, fcm_p, None,
            None, None,
            hel_l, spo_l, others,
            None, None, None,
        ]
        for col, val in enumerate(row_data2, 1):
            c = ws2.cell(row=cur2, column=col, value=val)
            c.font  = data_fnt; c.border = thin_brd
            if w_fill2:
                c.fill = w_fill2
            c.alignment = r_align if col not in (1, 2) else l_align
        cur2 += 1

    # ===================================================================
    # SHEET 3 – Scouting Summary  (plant-level pests)
    # ===================================================================
    ws3 = wb.create_sheet("Scouting Summary")
    for i, w in enumerate([22, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 32], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3["A1"] = "Production Site Name"; ws3["A1"].font = bold_fnt
    ws3["E1"] = "Year";                 ws3["E1"].font = bold_fnt
    ws3["F1"] = current_year
    ws3["A2"] = farm_title;             ws3["A2"].font = mk_font(bold=True, size=14)

    # Pest group headers (row 3)
    for rng, label, hex_c in [("C3:E3", "FCM", DARK_BLUE), ("F3:H3", "Helicoverpa", MID_BLUE), ("I3:K3", "Others", DARK_GREEN)]:
        ws3.merge_cells(rng)
        cell_ref = rng.split(":")[0]
        ws3[cell_ref] = label
        ws3[cell_ref].font    = hdr_fnt
        ws3[cell_ref].fill    = mk_fill(hex_c)
        ws3[cell_ref].alignment = c_align

    _h3 = [
        "Period", "GH No.",
        "Eggs", "Larvae", "Damages",
        "Eggs", "Larvae", "Damages",
        "Eggs", "Larvae", "Damages",
        "Remarks (Corrective action)",
    ]
    ws3.row_dimensions[3].height = 20
    ws3.row_dimensions[4].height = 20
    for col, hdr in enumerate(_h3, 1):
        c = ws3.cell(row=4, column=col, value=hdr)
        c.font = hdr_fnt; c.fill = dark_fill; c.border = thin_brd; c.alignment = c_align

    cur3 = 5
    for week_num in all_weeks:
        try:
            wk_mon, wk_sun = iso_week_range(current_year, week_num)
        except Exception:
            continue

        is_last = (week_num == last_week_num)
        w_fill3  = green_fill if is_last else None

        from_s = f"FROM: {human_date(wk_mon)}"
        to_s   = f"TO: {human_date(wk_sun)}"
        wk_lbl = f"Week {week_num:02d}"

        for idx, gh in enumerate(GHS):
            period = wk_lbl if idx == 0 else (from_s if idx == 1 else (to_s if idx == 2 else None))

            row_d = [
                period, gh,
                scouting_map.get((week_num, gh, "FCM",        "eggs"),    0),
                scouting_map.get((week_num, gh, "FCM",        "larvae"),  0),
                scouting_map.get((week_num, gh, "FCM",        "damages"), 0),
                scouting_map.get((week_num, gh, "Helicoverpa","eggs"),    0),
                scouting_map.get((week_num, gh, "Helicoverpa","larvae"),  0),
                scouting_map.get((week_num, gh, "Helicoverpa","damages"), 0),
                scouting_map.get((week_num, gh, "Others",     "eggs"),    0),
                scouting_map.get((week_num, gh, "Others",     "larvae"),  0),
                scouting_map.get((week_num, gh, "Others",     "damages"), 0),
                None,
            ]
            for col, val in enumerate(row_d, 1):
                c = ws3.cell(row=cur3, column=col, value=val)
                c.font  = data_fnt; c.border = thin_brd
                if w_fill3:
                    c.fill = w_fill3
                c.alignment = r_align if 3 <= col <= 11 else l_align
            cur3 += 1

    # ===================================================================
    # SHEET 4 – Intake QC Report  (intentionally empty — AI not integrated)
    # ===================================================================
    ws4 = wb.create_sheet("Intake QC Report")
    ws4.column_dimensions["A"].width = 22
    ws4["A1"] = "Production Site Name"; ws4["A1"].font = bold_fnt
    ws4["E1"] = "Year";                 ws4["E1"].font = bold_fnt
    ws4["F1"] = current_year
    ws4["A2"] = farm_title;             ws4["A2"].font = mk_font(bold=True, size=14)
    ws4["A4"] = "To be populated in the next update."
    ws4["A4"].font = italic_note_fnt

    # ===================================================================
    # SHEET 5 – Variety List  (intentionally empty — pending live integration)
    # ===================================================================
    ws5 = wb.create_sheet("Variety List")
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 36
    ws5["A1"] = "Production Site Name"; ws5["A1"].font = bold_fnt
    ws5["E1"] = "Year";                 ws5["E1"].font = bold_fnt
    ws5["F1"] = current_year
    ws5["A2"] = farm_title;             ws5["A2"].font = mk_font(bold=True, size=14)
    ws5["A4"] = "To be populated in the next update."
    ws5["A4"].font = italic_note_fnt

    # ===================================================================
    # SHEET 6 – FCM Risk profiling  (intentionally empty — AI not integrated)
    # ===================================================================
    ws6 = wb.create_sheet("FCM Risk profiling")
    ws6.column_dimensions["A"].width = 22
    ws6.column_dimensions["B"].width = 36
    ws6["A1"] = "Production Site Name"; ws6["A1"].font = bold_fnt
    ws6["C1"] = farm_title
    ws6["D1"] = "Month/Year";           ws6["D1"].font = bold_fnt
    ws6["E1"] = today.strftime("%B").upper()
    ws6["A4"] = "To be populated in the next update."
    ws6["A4"].font = italic_note_fnt

    # ===================================================================
    # Save workbook → bytes buffer
    # ===================================================================
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    fname_prefix = (abbrev or farm_display or "FCM").replace(" ", "_").upper()
    fname = f"{fname_prefix}_FCM_Weekly_Report_{current_year}_W{last_week_num:02d}.xlsx"

    return excel_bytes, fname, current_year, week_range_str, last_week_num, farm_display


# ---------------------------------------------------------------------------
# Email helpers
# ---------------------------------------------------------------------------

def _recipients():
    """FCM weekly Excel report recipients — resolved by role (no hardcoding)."""
    from upande_scp.serverscripts.reports.report_recipients import report_recipients

    return report_recipients()


def _build_email_html(week_range_str, current_year, farm_display, kephis_id):
    site_line = farm_display or "-"
    if kephis_id:
        site_line = f"{farm_display} ({kephis_id})"
    return f"""
    <div style="font-family:Arial,sans-serif;font-size:14px;color:#333;max-width:640px;">
      <div style="background:#0D2B5E;padding:22px 28px;margin-bottom:20px;border-radius:4px;">
        <div style="font-size:20px;font-weight:700;color:#fff;">FCM Weekly Monitoring Report</div>
        <div style="font-size:13px;color:#c8d8f0;margin-top:4px;">
          KEPHIS Site: {site_line} &mdash; {week_range_str} &mdash; {current_year}
        </div>
      </div>
      <p>Hi Team,</p>
      <p>Please find attached the <strong>KEPHIS FCM Weekly Monitoring Report</strong>
         for <strong>{farm_display}</strong>, covering <strong>{week_range_str}</strong>.</p>
      <table style="width:100%;border-collapse:collapse;font-size:13px;margin:12px 0;">
        <tr style="background:#0D2B5E;color:#fff;">
          <th style="padding:8px 12px;text-align:left;">Sheet</th>
          <th style="padding:8px 12px;text-align:left;">Contents</th>
        </tr>
        <tr style="background:#f4f6fb;">
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">FCM Daily monitoring</td>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">Per-trap FCM counts (indoor &amp; outdoor) by week</td>
        </tr>
        <tr>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">Weekly summary</td>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">FCM, Helicoverpa, Spodoptera, Duponchella &amp; Unidentified Moths by week</td>
        </tr>
        <tr style="background:#f4f6fb;">
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">Scouting Summary</td>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">Plant-level pest observations (eggs / larvae / damages) per GH</td>
        </tr>
        <tr>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">Intake QC Report</td>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;"><em>To be populated in the next update</em></td>
        </tr>
        <tr style="background:#f4f6fb;">
          <td style="padding:7px 12px;border-bottom:1px solid #dde;">Variety List</td>
          <td style="padding:7px 12px;border-bottom:1px solid #dde;"><em>To be populated in the next update</em></td>
        </tr>
        <tr style="background:#f4f6fb;">
          <td style="padding:7px 12px;">FCM Risk profiling</td>
          <td style="padding:7px 12px;"><em>To be populated in the next update</em></td>
        </tr>
      </table>
      <p style="font-size:13px;color:#555;">
        <strong>Last week&rsquo;s rows are highlighted in green</strong> throughout all data sheets.
        To trace any summary number back to the scout and scouting entry that produced
        it, download <strong>Scouting Entries</strong> for the same week from the Reports
        page — it lists every observation behind these totals.
      </p>
      <p>Regards,<br>Upande Crop-Protection System</p>
      <hr style="border:none;border-top:1px solid #ddd;margin:20px 0;">
      <p style="font-size:11px;color:#999;">
        Report generated: {frappe.utils.now()}<br>
        Site: {site_line}<br>
        Data covers all scouting entries for {current_year}.
      </p>
    </div>
    """


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

def send_fcm_weekly_excel_report():
    """Scheduler entry point — one workbook emailed per farm with data."""
    current_year = date.today().year
    farms = _farms_with_data(current_year)

    if not farms:
        _send_no_data_email(current_year, _week_range_str_now(), farm_display=None)
        return

    recipients = _recipients()
    sent = 0
    for farm in farms:
        excel_bytes, fname, _, week_range_str, _, farm_display = _build_workbook_bytes(farm)
        if excel_bytes is None:
            continue
        _, kephis_id, _ = _resolve_farm(farm)
        frappe.sendmail(
            recipients=recipients,
            subject=f"FCM Weekly Monitoring Report — {farm_display} — {week_range_str} ({current_year})",
            message=_build_email_html(week_range_str, current_year, farm_display, kephis_id),
            attachments=[{"fname": fname, "fcontent": excel_bytes}],
        )
        sent += 1
        frappe.logger().info(
            f"FCM Weekly Excel Report sent for {farm_display} ({week_range_str}) → {recipients}"
        )

    if sent == 0:
        _send_no_data_email(current_year, _week_range_str_now(), farm_display=None)


@frappe.whitelist()
def trigger_fcm_email(farm=None, week=None):
    """On-demand 'Send now' trigger from the Scouting Reports page.

    When ``farm`` is provided the email is limited to that farm; otherwise
    every farm with data is sent (same behaviour as the scheduler). ``week``
    optionally selects which ISO week the report targets.
    """
    if farm:
        current_year = date.today().year
        excel_bytes, fname, _, week_range_str, _, farm_display = _build_workbook_bytes(farm, week=week)
        if excel_bytes is None:
            _send_no_data_email(current_year, week_range_str, farm_display=farm_display)
            return {"ok": False, "farm": farm_display, "reason": "no data"}
        _, kephis_id, _ = _resolve_farm(farm)
        frappe.sendmail(
            recipients=_recipients(),
            subject=f"FCM Weekly Monitoring Report — {farm_display} — {week_range_str} ({current_year})",
            message=_build_email_html(week_range_str, current_year, farm_display, kephis_id),
            attachments=[{"fname": fname, "fcontent": excel_bytes}],
        )
        return {"ok": True, "farm": farm_display, "recipients": _recipients()}

    send_fcm_weekly_excel_report()
    return {"ok": True, "recipients": _recipients()}


def save_fcm_xlsx_to_path(path, farm):
    """Build the report for ``farm`` and write it to ``path`` on disk.

    Intended for ad-hoc use (e.g. ``bench execute`` to generate a sample for
    doc references). Returns the filename written, or raises if no data.
    """
    import os

    excel_bytes, fname, current_year, _, _, farm_display = _build_workbook_bytes(farm)
    if excel_bytes is None:
        frappe.throw(
            f"No scouting data found for {farm_display or farm} in {current_year}. "
            f"Report not generated."
        )

    target = os.path.join(path, fname) if os.path.isdir(path) else path
    with open(target, "wb") as f:
        f.write(excel_bytes)
    return target


@frappe.whitelist()
def download_fcm_xlsx(farm, week=None):
    """Stream the freshly-built xlsx back as a browser download.

    ``week`` optionally selects which ISO week the report targets; when
    omitted the last completed week is used.
    """
    if not farm:
        frappe.throw("A farm must be supplied to download the FCM weekly report.")
    excel_bytes, fname, current_year, _, _, farm_display = _build_workbook_bytes(farm, week=week)
    if excel_bytes is None:
        frappe.throw(
            f"No scouting data found for {farm_display or farm} in {current_year}. "
            f"Report not generated."
        )
    frappe.local.response.filename = fname
    frappe.local.response.filecontent = excel_bytes
    frappe.local.response.type = "download"


def _build_scouting_entries_bytes(farm, week=None, year=None):
    """Build the standalone Scouting Entries workbook for one farm and week.

    This is the per-observation audit trail that used to ride along inside the
    KEPHIS workbook as a 7th sheet. It was moved out for two reasons:

    * Cost. The sheet carried every observation from week 1 to the selected
      week, and styled all fifteen cells of every row individually. On a large
      farm late in the year that was ~350,000 rows and around two minutes of
      openpyxl — enough to blow past the request timeout and take the KEPHIS
      submission down with it.
    * Scope. The KEPHIS submission covers one week, so an audit trail for that
      week is what reconciles it. A year-to-date trail was answering a question
      nobody asked of a weekly report.

    So this builds ONE week, and styles the header row only — column widths and
    a freeze pane, nothing per-cell. Writing 15 cells with a font, a border and
    an alignment each costs roughly 13x what writing the values alone costs;
    at this size that is the difference between a snappy download and a slow
    one, and it buys nothing on a sheet people filter and pivot.

    Returns (excel_bytes, fname, current_year, week_num, farm_display).
    ``excel_bytes`` is None when the week holds no observations.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.log_error("openpyxl not installed – run: pip install openpyxl", "FCM Weekly Report")
        return None, None, None, None, None

    farm_display, _, abbrev = _resolve_farm(farm)

    today = date.today()
    current_year = int(year) if year else today.year
    default_week_mon = today - timedelta(days=today.weekday() + 7)
    week_num = int(week) if week else int(default_week_mon.isocalendar()[1])

    try:
        wk_mon = date.fromisocalendar(current_year, week_num, 1)
        wk_sun = date.fromisocalendar(current_year, week_num, 7)
        week_range_str = wk_mon.strftime("%d %b") + " - " + wk_sun.strftime("%d %b %Y")
    except ValueError:
        week_range_str = f"Week {week_num}"

    # The week is pushed into SQL rather than filtered here: this query returns
    # raw un-aggregated rows, so fetching a whole year to keep one week of it
    # would cost ~25s on a large farm and throw away 97% of what it read.
    # get_scouting_records_weekly de-duplicates re-submitted observations, so
    # these rows reconcile 1:1 with the KEPHIS workbook's summary sheets.
    records = get_scouting_records_weekly(current_year, farm=farm, week=week_num)
    if not records:
        return None, None, current_year, week_num, farm_display

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scouting Entries"

    headers = [
        "Week", "Date", "Time", "Scout",
        "Greenhouse", "Bed", "Zone", "Block", "Row", "Tree",
        "Type", "Trap / Location", "Pest", "Count", "Scouting Entry",
    ]
    widths = [8, 13, 10, 26, 30, 16, 16, 12, 12, 16, 12, 14, 18, 10, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append([f"{farm_display} — Scouting Entries — {week_range_str} ({current_year})"])
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.append([])
    ws.append(headers)

    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0D2B5E")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align

    def safe_int(v):
        try:
            return int(float(v or 0))
        except (TypeError, ValueError):
            return 0

    for rec in records:
        entry_type = rec.get("entry_type") or ""
        if entry_type == "Trap":
            trap_loc = " — ".join(
                x for x in (rec.get("trap") or "", rec.get("stage_or_location") or "") if x
            )
            pest_lbl = rec.get("pest") or ""
        else:
            trap_loc = ""
            stage_lbl = rec.get("stage_or_location") or ""
            pest_lbl = rec.get("pest") or ""
            if stage_lbl:
                pest_lbl = f"{pest_lbl} ({stage_lbl})" if pest_lbl else stage_lbl

        # ws.append() writes the whole row in one call and leaves every cell
        # on the workbook default style — this is the cheap path.
        ws.append([
            f"WK {rec.get('wk'):02d}" if rec.get("wk") is not None else "",
            rec.get("date_of_capture"),
            str(rec.get("time_of_capture") or ""),
            rec.get("scout") or "",
            rec.get("greenhouse") or "",
            rec.get("bed") or "",
            rec.get("zone") or "",
            rec.get("block") or "",
            rec.get("row") or "",
            rec.get("tree") or "",
            entry_type,
            trap_loc,
            pest_lbl,
            safe_int(rec.get("count")),
            rec.get("scouting_entry") or "",
        ])

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    prefix = (abbrev or farm_display or "FCM").replace(" ", "_").upper()
    fname = f"{prefix}_Scouting_Entries_{current_year}_W{week_num:02d}.xlsx"
    return buf.read(), fname, current_year, week_num, farm_display


@frappe.whitelist()
def download_scouting_entries_xlsx(farm, week=None):
    """Stream the per-observation Scouting Entries workbook for one week."""
    if not farm:
        frappe.throw("A farm must be supplied to download scouting entries.")
    excel_bytes, fname, current_year, week_num, farm_display = (
        _build_scouting_entries_bytes(farm, week=week)
    )
    if excel_bytes is None:
        frappe.throw(
            f"No scouting entries found for {farm_display or farm} "
            f"in week {week_num} of {current_year}."
        )
    frappe.local.response.filename = fname
    frappe.local.response.filecontent = excel_bytes
    frappe.local.response.type = "download"


@frappe.whitelist()
def list_farms_with_data(year=None):
    """Return farms that have scouting data for the given (or current) year.

    Used by the Scouting Reports page to populate a farm selector before
    triggering a send/download.
    """
    current_year = int(year) if year else date.today().year
    farms = _farms_with_data(current_year)
    out = []
    for f in farms:
        display, kephis, abbrev = _resolve_farm(f)
        out.append({
            "farm": f,
            "display": display,
            "kephis_farm_id": kephis,
            "abbreviation": abbrev,
        })
    return out


@frappe.whitelist()
def list_report_weeks(farm, year=None):
    """Return the ISO weeks (Mon → Sun) that have scouting data for ``farm``.

    Powers the week selector on the Reports page. Weeks are returned newest
    first, capped at the current week, each with a human label and the
    Monday/Sunday bounds.
    """
    if not farm:
        return []
    current_year = int(year) if year else date.today().year
    current_week = int(date.today().isocalendar()[1])

    rows = frappe.db.sql(
        """
        SELECT DISTINCT WEEK(se.date_of_capture, 1) AS wk
        FROM  `tabScouting Entry` se
        JOIN  `tabWarehouse` gh ON gh.name = se.greenhouse
        WHERE YEAR(se.date_of_capture) = %s
          AND gh.custom_farm = %s
        ORDER BY wk DESC
        """,
        (current_year, farm),
        as_dict=True,
    )

    out = []
    for r in rows:
        wk = r.wk
        if wk is None or wk < 1 or wk > current_week:
            continue
        try:
            mon = date.fromisocalendar(current_year, wk, 1)
            sun = date.fromisocalendar(current_year, wk, 7)
        except ValueError:
            continue
        out.append({
            "week": wk,
            "label": f"Week {wk:02d} ({mon.strftime('%d %b')} – {sun.strftime('%d %b')})",
            "from": mon.isoformat(),
            "to": sun.isoformat(),
        })
    return out


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _week_range_str_now():
    today = date.today()
    last_week_mon = today - timedelta(days=today.weekday() + 7)
    last_week_sun = last_week_mon + timedelta(days=6)
    return last_week_mon.strftime("%d %b") + " - " + last_week_sun.strftime("%d %b %Y")


def _send_no_data_email(current_year, week_range_str, farm_display=None):
    """Send a brief notice when no scouting data exists."""
    recipients = _recipients()

    who = f" for {farm_display}" if farm_display else ""
    frappe.sendmail(
        recipients=recipients,
        subject=f"FCM Weekly Report — No Data{who} ({week_range_str})",
        message=(
            f"<p>Hi Team,</p>"
            f"<p>No scouting data found{who} for {current_year}. "
            f"The FCM weekly Excel report was not generated.</p>"
            f"<p>Regards,<br>Upande System</p>"
        ),
    )
