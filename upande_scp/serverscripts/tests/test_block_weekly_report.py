"""The weekly block sheet for avocado and coffee."""

import io
import unittest

import frappe

from upande_scp.serverscripts.reports import block_weekly_report as R


class TestAvailability(unittest.TestCase):
	def test_avocado_is_ready_because_lokitela_has_blocks(self):
		frappe.set_user("Administrator")
		a = R.availability("Avocado")
		farms = [r["farm"] for r in a["ready"]]
		self.assertIn("Lokitela", farms)

	def test_coffee_is_blocked_and_says_why(self):
		"""Endebess and Saboti are tagged for coffee but have no warehouse typed as a
		block, so a download would be blank for a reason that looks like a clean week."""
		frappe.set_user("Administrator")
		a = R.availability("Coffee")
		self.assertEqual(a["ready"], [])
		blocked = {r["farm"]: r["reason"] for r in a["blocked"]}
		self.assertIn("Endebess", blocked)
		self.assertIn("Saboti", blocked)
		self.assertIn("no blocks", blocked["Endebess"].lower())

	def test_a_farm_with_no_blocks_refuses_to_build(self):
		frappe.set_user("Administrator")
		with self.assertRaises(frappe.ValidationError):
			R.build_workbook_bytes("Coffee", "Endebess", 2026, 28)


class TestTheSheet(unittest.TestCase):
	YEAR, WEEK = 2026, 28

	def _sheet(self):
		from openpyxl import load_workbook

		frappe.set_user("Administrator")
		data = R.build_workbook_bytes("Avocado", "Lokitela", self.YEAR, self.WEEK)
		return load_workbook(io.BytesIO(data)).active

	def test_every_block_gets_a_row_even_with_no_findings(self):
		"""A block walked and found clean, and a block nobody visited, are different
		facts. Omitting both makes them identical."""
		ws = self._sheet()
		blocks = R._blocks_for_farm("Lokitela")
		self.assertTrue(blocks)
		names = set()
		for r in range(6, ws.max_row):
			v = ws.cell(row=r, column=1).value
			if v:
				names.add(v)
		for b in blocks:
			self.assertIn(b, names)

	def test_pests_are_columns_and_cells_are_counts(self):
		ws = self._sheet()
		header = []
		for c in range(1, ws.max_column + 1):
			header.append(ws.cell(row=5, column=c).value)
		self.assertEqual(header[0], "Block")
		self.assertEqual(header[-1], "Total")
		self.assertGreater(len(header), 2, "expected at least one pest column")

	def test_the_totals_row_adds_up(self):
		ws = self._sheet()
		total_row = ws.max_row
		self.assertEqual(ws.cell(row=total_row, column=1).value, "Total")
		# Column totals must equal the sum of the rows above them.
		for c in range(2, ws.max_column):
			column_sum = 0.0
			for r in range(6, total_row):
				column_sum += float(ws.cell(row=r, column=c).value or 0)
			self.assertAlmostEqual(
				float(ws.cell(row=total_row, column=c).value or 0), column_sum
			)

	def test_the_week_is_named_in_the_sheet(self):
		ws = self._sheet()
		self.assertIn(f"{self.YEAR}-W{self.WEEK:02d}", str(ws["A2"].value))


class TestWeeks(unittest.TestCase):
	def test_only_weeks_with_scouting_are_offered(self):
		frappe.set_user("Administrator")
		weeks = R.report_weeks("Avocado", "Lokitela")
		self.assertTrue(weeks)
		for w in weeks:
			self.assertGreater(w["entries"], 0)

	def test_a_farm_with_no_blocks_offers_no_weeks(self):
		frappe.set_user("Administrator")
		self.assertEqual(R.report_weeks("Coffee", "Endebess"), [])


class TestBlocksReachThePlanner(unittest.TestCase):
	"""Blocks have to be offered to the planner, and with their area.

	`_resolve_user_scope` splits warehouses into `greenhouses` by type, and the
	bootstrap then runs those through `is_greenhouse_allowed` — a filter that enforces
	the `GH <N>` naming convention and the tunnel/phase/ipm/csu exclusions. Every one
	of those rules rejects `AIRSTRIP BLK 10 - KL`, so blocks need their own path.
	"""

	def test_blocks_carry_their_area(self):
		from upande_scp.serverscripts.spray_plan_creator.bootstrap import _enrich_blocks

		frappe.set_user("Administrator")
		raw = frappe.get_all(
			"Warehouse",
			filters={"custom_farm": "Lokitela", "warehouse_type": "Block", "disabled": 0},
			fields=["name", "custom_farm", "warehouse_type"],
		)
		if not raw:
			self.skipTest("no blocks on Lokitela")
		blocks = _enrich_blocks(raw)
		self.assertEqual(len(blocks), len(raw))
		with_area = [b for b in blocks if b["area_ha"] > 0]
		self.assertTrue(with_area, "no block carried an area — quantities cannot derive")
		for b in blocks:
			self.assertIn("custom_farm", b)
			self.assertIn("cost_center", b)

	def test_the_rose_area_source_is_empty_for_avocado(self):
		"""Why blocks need `custom_area_ha` at all: summing units the rose way is zero."""
		frappe.set_user("Administrator")
		total = frappe.db.sql(
			"""SELECT IFNULL(SUM(b.bed__area), 0) FROM `tabBed` b
			   JOIN `tabWarehouse` w ON w.name = b.greenhouse
			   WHERE w.custom_farm = 'Lokitela'"""
		)[0][0]
		self.assertEqual(float(total or 0), 0.0)
